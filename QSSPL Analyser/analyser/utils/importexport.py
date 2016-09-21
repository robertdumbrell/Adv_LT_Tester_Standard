
import numpy as np
import os
import sys
from shutil import copyfile
import json
import openpyxl as pyxl
import scipy.constants as C


class Load_sinton():

    def __init__(self, Directory, RawDataFile):
        self.Directory = Directory
        self.RawDataFile = RawDataFile

    def Load_RawData_File(self):
        '''
        Loads a Sinton excel and passes it into a lifetime class, with the
        attributes automatically filled. You still need to check that the Sinton
        excel values were correctly choosen.
        '''

        # define the lifetime class
        # get the measurement data
        file_path = os.path.join(self.Directory, self.RawDataFile)

        wb = pyxl.load_workbook(file_path, read_only=True, data_only=True)
        data = self._openpyxl_Sinton2014_ExtractRawDatadata(wb)
        inf = self._openpylx_sinton2014_extractsserdata(wb)

        data.dtype.names = ('Time', 'PC', 'Gen', 'PL')
        data['PC'] += inf['dark_voltage']

        return data

    def _openpyxl_Sinton2014_ExtractRawDatadata(self, wb):
        '''
            reads the raw data a sinton WCT-120 spreadsheet form the
            provided instance of the openpylx workbook.
        '''

        # make sure the sheet is in the book
        assert 'Calc' in wb.get_sheet_names()

        # get the worksheet
        ws = wb.get_sheet_by_name('Calc')

        # get first section of data
        values1 = np.array([[i.value for i in j] for j in ws['A9':'C133']],
                           dtype=np.float64)

        # add in values so that the background correction works
        for repeat in range(int(values1.shape[0] * 0.1)):
            values1 = np.vstack((values1, values1[-1, :]))
            values1[2, -1] = 0

        headers1 = tuple(
            [[j.value for j in i] for i in ws['A8':'C8']][0])

        # get second section of data
        values2 = np.zeros(values1.shape[0])
        headers2 = ('PL',)

        # form into one array with names
        values = np.vstack((values1.T, values2)).T
        headers = headers1 + headers2

        dtype = [('a', np.float64), ('b', np.float64),
                 ('c', np.float64), ('d', np.float64)]

        Out = values.copy().view(dtype=dtype).reshape(values.shape[0],)

        time_diff = Out['a'][2] - Out['a'][1]
        for i in range(Out['a'].shape[0]):
            Out['a'][i] = i * time_diff - time_diff

        return Out

    def _openpylx_sinton2014_extractsserdata(self, wb):

        # make sure the sheet is in the book
        # get the worksheet
        assert 'User' in wb.get_sheet_names()
        assert 'Settings' in wb.get_sheet_names()

        ws = wb.get_sheet_by_name('User')

        # Grabbing the data and assigning it a nae

        user_set = {
            'Thickness': float(ws['B6'].value),
            'Doping': float(ws['J9'].value),
            'sample_type': ws['D6'].value.encode('utf8'),
            'optical_constant': float(ws['E6'].value),
        }

        user_set['Reflection'] = (1 - user_set['optical_constant']) * 100

        # makes a reference to the RawData page
        ws = wb.get_sheet_by_name('Settings')

        # grabs the Cell ref and terns it into a FS
        sys_set = {
            'Fs': 0.038 / C.e / float(ws['C5'].value),
        }

        # make one dic
        user_set.update(sys_set)

        ws = wb.get_sheet_by_name('Calc')
        sys_set = {
            'dark_voltage': float(ws['B166'].value),
        }

        user_set.update(sys_set)

        return user_set

    def Load_InfData_File(self):
        # print 'Still under construction'

        # replace the ending with a new ending
        InfFile = self.get_inf_name()
        InfFile = os.path.join(self.Directory, InfFile)
        # These are adjustment Values, requried by the following
        temp_list = {'Doping': 1,
                     'Thickness': 1,
                     'Binning': 1,
                     'Reflection': 0.0,
                     'Fs': 1,
                     'Ai': 1,
                     'Quad': 0.0004338,
                     'Lin': 0.03611,
                     'Const': 0,
                     'CropStart': None,
                     'CropEnd': None,
                     'Waveform': 'blank',
                     'Temp': 300,
                     }

        if os.path.isfile(InfFile):
            with open(InfFile, 'r') as f:
                file_contents = f.read()
                List = json.loads(file_contents)
        else:
            file_path = os.path.join(self.Directory, self.RawDataFile)
            wb = pyxl.load_workbook(file_path, read_only=True, data_only=True)
            List = self._openpylx_sinton2014_extractsserdata(wb)
            temp_list.update(List)
            serialised_json = json.dumps(
                temp_list,
                sort_keys=True,
                indent=4,
                separators=(',', ': ')
            )

            # writes to file
            with open(InfFile, 'w+') as text_file:
                text_file.write(serialised_json)

        temp_list.update(List)

        return temp_list

    def Load_ProcessedData_File(self):
        DataFile = self.DataFile[:-13] + '.dat'
        return np.genfromtxt(self.Directory + DataFile, usecols=(0, 1, 8, 9), unpack=True, delimiter='\t', names=('Deltan_PC', 'Tau_PC', 'Deltan_PL', 'Tau_PL'))

    def WriteTo_Inf_File(self, metadata_dict):
        '''
        write to inf file, with "correct format" format as
        '''

        # get inf file location
        InfFile = os.path.join(self.Directory, self.get_inf_name())

        # check if there is backup, and make
        backup_file = os.path.join(self.Directory, InfFile + ".Backup")
        if os.path.isfile(backup_file) is False:

            copyfile(InfFile, backup_file)
            print 'Backuped original .inf  file as .inf.backup'

        # specify how to output

        serialised_json = json.dumps(
            metadata_dict,
            sort_keys=True,
            indent=4,
            separators=(',', ': ')
        )

        # writes to file
        with open(InfFile, 'w') as text_file:
            text_file.write(serialised_json)

    def get_inf_name(self):

        if self.RawDataFile.count('.xlsm') == 1:
            InfFile = self.RawDataFile.replace('.xlsm', '.json')
        else:
            print 'stop fucking around with the name!!'

        return InfFile


class Load_QSSPL_File_LabView():

    def __init__(self, Directory, RawDataFile):
        self.Directory = Directory
        self.RawDataFile = RawDataFile

    def Load_RawData_File(self):
        return np.genfromtxt(os.path.join(self.Directory, self.RawDataFile),
                             names=('Time', 'PC', 'Gen', 'PL'))

    def Load_InfData_File(self):
        InfFile = self.RawDataFile[:-13] + '.inf'

        '''info from inf file '''

        Cycles, dump, Frequency, LED_Voltage, dump, dump, dump, dump, DataPoints, dump = np.genfromtxt(
            self.Directory + InfFile, skip_header=20, skip_footer=22, delimiter=':', usecols=(1), autostrip=True, unpack=True)
        Waveform, LED_intensity = np.genfromtxt(
            self.Directory + InfFile, skip_header=31, skip_footer=20, delimiter=':', usecols=(1), dtype=None, autostrip=True, unpack=True)

        l = np.genfromtxt(
            self.Directory + InfFile, skip_header=36, delimiter=':', usecols=(1))

        Doping = l[9]
        Ai = l[6]
        Fs = l[7]
        Thickness = l[12]
        Quad = l[12]
        Lin = l[12]
        Const = 0

        CropStart = None
        CropEnd = None

        Binning = int(l[2])
        Reflection = (1 - l[16]) * 100

        dic = locals()

        del dic['self']
        del dic['l']
        del dic['dump']

        return dic

    def Load_ProcessedData_File(self):
        DataFile = self.DataFile[:-13] + '.dat'
        return np.genfromtxt(self.Directory + DataFile, usecols=(0, 1, 8, 9), unpack=True, delimiter='\t', names=('Deltan_PC', 'Tau_PC', 'Deltan_PL', 'Tau_PL'))

    def WriteTo_Inf_File(self, Dictionary):

        InfFile = self.RawDataFile[:-13] + '.inf'

        if (os.path.isfile(self.Directory + InfFile + ".Backup") == False):
            copyfile(
                self.Directory + InfFile, self.Directory + InfFile + ".Backup")
            print 'Backuped original .inf  file as .inf.backup'

        ####
        # Creating the .inf file, this can be done more easily with list(f), but i'm not using it right now.
        ###
        f = open(self.Directory + InfFile, 'r')

        # print list(f)
        # print list(f).shape

        s = ''
        for i in range(38):
            s = s + f.readline()
        s = s + f.readline()[:26] + \
            '{0:.0f}'.format(Dictionary['Binning']) + '\n'
        for i in range(3):
            s = s + f.readline()
        s = s + f.readline()[:5] + '{0:.3e}'.format(Dictionary['Ai']) + '\n'
        s = s + f.readline()[:11] + '{0:.3e}'.format(Dictionary['Fs']) + '\n'
        s = s + f.readline()
        s = s + f.readline()[:23] + \
            '{0:.3e}'.format(Dictionary['Doping']) + '\n'
        s = s + f.readline()
        s = s + f.readline()
        s = s + f.readline()[:12] + \
            '{0:.4f}'.format(Dictionary['Thickness']) + '\n'
        s = s + f.readline()[:24] + '{0:.4e}'.format(Dictionary['Quad']) + '\n'
        s = s + f.readline()[:21] + '{0:.4e}'.format(Dictionary['Lin']) + '\n'
        s = s + f.readline()
        s = s + \
            f.readline()[
                :37] + '{0:.6f}'.format(1 - Dictionary['Reflection'] / 100) + '\n'

        for i in range(6):
            s = s + f.readline()

        f.close()
        f = open(self.Directory + InfFile, 'w')
        f.write(s)


class Load_QSSPL_File_Python():

    def __init__(self, Directory, RawDataFile):
        self.Directory = Directory
        self.RawDataFile = RawDataFile

    def Load_RawData_File(self):
        data = np.genfromtxt(
            self.Directory + self.RawDataFile, unpack=True, names=True, delimiter='\t')
        s = np.array([])
        dic = {'Time_s': 'Time', 'Gen_V': 'Gen', 'Generation_V': 'Gen',
               'PL_V': 'PL', 'PC_V': 'PC'}
        # print np.array(data.dtype.names)
        for i in np.array(data.dtype.names):
            # print i,dic[i]
            s = np.append(s, dic[i])

        # print s

        data.dtype.names = s
        # ('Time','Gen','PL','PC')
        return data

    def num(self, s):
        try:
            return float(s)
        except ValueError:
            return s

    def Load_InfData_File(self):
        # print 'Still under construction'

        InfFile = self.RawDataFile[:-13] + '.inf'

        # These are adjustment Values
        Doping = 1
        Thickness = 1
        Binning = 1
        Reflection = 0.0
        Fs = 1
        Ai = 1
        Quad = 0.0004338
        Lin = 0.03611
        Const = 0.001440789
        Temp = 300

        Waveform = None
        CropStart = None
        CropEnd = None

        List = locals()

        del List['InfFile']
        del List['self']

        with open(self.Directory + str(InfFile), 'r') as f:
            s = f.read()

        s = s.replace('\n\n', '\n')
        for i in s.split('\n')[2:-1]:
            try:
                List[i.split(':\t')[0].strip()] = self.num(i.split(':\t')[1])
            except:
                List[i.split('\t')[0].strip()] = self.num(i.split('\t')[1])

        return List

    def Load_ProcessedData_File(self):
        print 'Still under construction'

        return zeros(4, 4)

    def WriteTo_Inf_File(self, Dictionary):

        InfFile = self.RawDataFile[:-13] + '.inf'

        if (os.path.isfile(self.Directory + InfFile + ".Backup") == False):
            copyfile(
                self.Directory + InfFile, self.Directory + InfFile + ".Backup")
            print 'Backuped original .inf  file as .inf.backup'

        s = 'MJ system\r\nList of vaiables:\r\n'
        for i in Dictionary:
            # if i != and i !='self':
            s += '{0}:\t{1}\r\n'.format(i, Dictionary[i])

        with open(self.Directory + InfFile, 'w') as text_file:
            text_file.write(s)


class TempDep_loads():

    def __init__(self, Directory, RawDataFile):
        self.Directory = Directory
        self.RawDataFile = RawDataFile

    def Load_RawData_File(self):
        '''
        Loads the measured data from the data file.
        This has the file extension tsv (tab seperated values)

        from a provided file name,
        takes data and outputs data with specific column headers
        '''

        # get data, something stange was happening with os.path.join
        file_location = os.path.normpath(
            os.path.join(self.Directory, self.RawDataFile))

        data = np.genfromtxt(
            os.path.join(file_location),
            unpack=True, names=True, delimiter='\t')

        # string to convert file names to program names
        dic = {'Time_s': 'Time', 'Generation_V': 'Gen',
               'PL_V': 'PL', 'PC_V': 'PC'}

        # create empty array
        s = np.array([])

        # build array of names, in correct order
        for i in np.array(data.dtype.names):
            s = np.append(s, dic[i])

        # assign names
        data.dtype.names = s

        return data

    def num(self, s):
        '''
        converts s to a number, or returns s
        '''
        try:
            return float(s)
        except ValueError:
            return s

    def Load_InfData_File(self):
        # print 'Still under construction'

        # replace the ending with a new ending
        InfFile = self.get_inf_name()

        # These are adjustment Values, requried by the following
        temp_list = {'Doping': 1,
                     'Thickness': 1,
                     'Binning': 1,
                     'Reflection': 0.0,
                     'Fs': 1,
                     'Ai': 1,
                     'Quad': 0.0004338,
                     'Lin': 0.03611,
                     'CropStart': None,
                     'CropEnd': None,
                     }

        with open(os.path.join(self.Directory, InfFile), 'r') as f:
            file_contents = f.read()
            List = json.loads(file_contents)

        List.update(temp_list)

        return List

    def Load_ProcessedData_File(self):
        print 'Still under construction'

        return zeros(4, 4)

    def WriteTo_Inf_File(self, metadata_dict):
        '''
        write to inf file, with "correct format" format as
        '''

        # get inf file location
        InfFile = os.path.join(self.Directory, self.get_inf_name())

        # check if there is backup, and make
        backup_file = os.path.join(self.Directory, InfFile + ".Backup")
        if os.path.isfile(backup_file) is False:

            copyfile(InfFile, backup_file)
            print 'Backuped original .inf  file as .inf.backup'

        # specify how to output

        serialised_json = json.dumps(
            metadata_dict,
            sort_keys=True,
            indent=4,
            separators=(',', ': ')
        )

        # writes to file
        with open(InfFile, 'w') as text_file:
            text_file.write(serialised_json)

    def get_inf_name(self):

        if self.RawDataFile.count('.tsv') == 1:
            InfFile = self.RawDataFile.replace('.tsv', '.json')
        else:
            print 'stop fucking around with the name!!'

        return InfFile


class LoadData():

    Directory = ''
    RawDataFile = ''
    File_Type = ''

    file_ext_dic = {
        '.Raw Data.dat': 'Python',
        '_Raw Data.dat': 'Labview',
        '.tsv': 'TempDep',
        '.xlsm': 'sinton'
    }
    file_ext_2_class = {
        r'.Raw Data.dat': r'Load_QSSPL_File_Python',
        r'_Raw Data.dat': r'Load_QSSPL_File_LabView',
        r'.tsv': r'TempDep_loads',
        r'.xlsm': r'Load_sinton',
    }

    def obtain_operatorclass(self, Directory=None, RawDataFile=None):
        if Directory is None:
            Directory = self.Directory
            RawDataFile = self.RawDataFile
        LoadClass = None

        for ext, _class in self.file_ext_2_class.iteritems():
            # print ext, _class

            if ext in RawDataFile:
                LoadClass = getattr(sys.modules[__name__],
                                    _class)(
                    Directory, RawDataFile)

        return LoadClass

    def Load_RawData_File(self):
        LoadClass = self.obtain_operatorclass()
        return LoadClass.Load_RawData_File()

    def Load_InfData_File(self):
        LoadClass = self.obtain_operatorclass()
        return LoadClass.Load_InfData_File()

    def Load_ProcessedData_File(self):
        LoadClass = self.obtain_operatorclass()
        return LoadClass.Load_ProcessedData_File()

    def WriteTo_Inf_File(self, Dict):
        LoadClass = self.obtain_operatorclass()
        return LoadClass.WriteTo_Inf_File(Dict)

if __name__ == "__main__":
    Folder = r'C:\git\ui\pvapp\test\data'
    File = r'raw_test_data.tsv'
    LoadData()
    B = LoadData().obtain_operatorclass(Folder, File)

    dictr = B.Load_InfData_File()
    B.WriteTo_Inf_File(dictr)
